
import CosmosDB
import datetime 
import pandas as pd
from openpyxl import load_workbook

wb = load_workbook("後期授業予定.xlsx")
ws = wb["R5"]

date_ln = 1
for i in range(191):
    if ws['d'+str(i+1)].value != None:
        date = ws["a"+str(i+1)].value
        print(date)
        if date == None:
            date = ws["a"+str(date_ln)].value
        else:
            date_ln = i+1
        date = pd.to_datetime(date)
        date_str = str(date.month) + "-" + str(date.day)
        print(date_str)

        Schedule_cell = []
        for j in [ws[j+str(i+1)].value if ws[j+str(i+1)].value != None else "" for j in ["e", "f", "g", "h", "i"]]:Schedule_cell.append(j)

        Description = ws['c'+str(date_ln)].value
        Note = ws['j'+str(date_ln)].value

        for j in ["G"+j for j in str(ws['d'+str(i+1)].value)]:
            print(j)
            CosmosDB.WriteDBItem(j, { "id" : date_str,  "Timetable" : Schedule_cell, "Events" : Description if Description != None else "", "Notes" : Note if Note != None else "" })